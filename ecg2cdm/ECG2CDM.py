import numpy as np
import pandas as pd
from openpyxl import load_workbook
import re
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity

class EKG_rule:
    def __init__(self):
        self.Rule = None
        self.should_not_use = None
        self.comment3 = None
        self.comment4 = None
        self.X = None
        self.y = None
        self.tfidf = None
        self.tfidf_matrix = None
        self.index_list = None
        self.idx_max = None

    def load_data(self, filename, additional=False):
        load_wb = load_workbook(filename, data_only=True)
        load_ws = load_wb.active
        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
        header = data.iloc[0]
        data = data[1:]
        data.rename(columns=header, inplace=True)
        data.reset_index(drop=True, inplace=True)

        if not additional:
            self.idx_max = len(data)
        else:
            a_len = self.idx_max
            self.idx_max = a_len + len(data)
            data.index = range(a_len, a_len+len(data))

        self.Rule = data[['source_name', 'condition_concept_id', 'concept_name']]
        self.X, self.y = self.process_data(data, additional)

        self.tfidf = CountVectorizer(stop_words=['consider', 'probable', 'probably', 'possible', 'suggests', 'prob'], token_pattern=r"\b\S\S+\b")
        self.tfidf_matrix = self.tfidf.fit_transform(self.X)

        self.should_not_use = list(self.X[data['should_not_use'] == 2])
        self.comment3 = list(self.X[data['comment'] == 3])
        self.comment4 = list(self.X[data['comment'] == 4])

    def process_data(self, data, additional):
        star = self.Rule['source_name'].str.replace(pat=r'\*', repl='', regex=False)
        Newline = star.str.replace(pat=r'\{New Line\}', repl=' ', regex=False)

        X = Newline.str.replace(pat='-', repl=' ', regex=False)
        X2 = Newline.str.replace(pat='-', repl='', regex=False)
        X = X.append(X2, ignore_index=False)
        X = X.str.lower()
        X.drop_duplicates(inplace=True)

        if additional:
            self.X = self.X.append(X, ignore_index=False)
            self.X.drop_duplicates(inplace=True)

        y = self.Rule[['condition_concept_id', 'concept_name']]
        if additional:
            self.y = self.y.append(y, ignore_index=False)

        self.index_list = self.X.index.tolist()

        return X, y


    def additional_Data_Load(self, filename, data_only=True):
        load_wb = load_workbook(filename, data_only)
        rule = []
        load_ws = load_wb.active  # first sheet
        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
        header = data.iloc[0]
        data = data[1:]
        data.rename(columns=header, inplace=True)
        a_len = self.idx_max
        self.idx_max = a_len + len(data)
        data.index = range(a_len, a_len+len(data))

        Rule = data[['source_name', 'condition_concept_id', 'concept_name']]

        star = Rule['source_name'].str.replace(pat=r'\*', repl='', regex=False)
        Newline = star.str.replace(pat=r'\{New Line\}', repl=' ', regex=False)

        X = Newline.str.replace(pat='-', repl=' ', regex=False)
        X2 = Newline.str.replace(pat='-', repl='', regex=False)
        X = X.append(X2, ignore_index=False)
        X = X.str.lower()
        X.drop_duplicates(inplace=True)

        self.X = self.X.append(X, ignore_index=False)
        self.X.drop_duplicates(inplace=True)

        y = Rule[['condition_concept_id', 'concept_name']]
        self.y = self.y.append(y, ignore_index=False)

        self.index_list = self.X.index.tolist()

        #self.tfidf = CountVectorizer(stop_words=['consider', 'probable', 'probably', 'possible', 'suggests', 'prob'], tokenizer=word_tokenize)

        self.tfidf_matrix = self.tfidf.fit_transform(self.X)

        if (data['should_not_use']==2).any(): self.should_not_use.append(list(Newline[data['should_not_use'] == 2].str.lower()))
        if (data['comment'] == 3).any(): self.comment3.append(list(Newline[data['comment'] == 3].str.lower()))
        if (data['comment'] == 4).any(): self.comment4.append(list(Newline[data['comment'] == 4].str.lower()))

    def OrderedSet(self, list):
        my_set = set()
        res = []
        for e in list:
            if e not in my_set:
                res.append(e)
                my_set.add(e)
        return res

    def get_similar(self, statement=None):
        if statement == None: statement = list(self.X)
        concept_id = []
        concept_name = []
        #mapping_text = []
        for input_ in statement:
            input_ = re.sub('-', ' ', input_)
            input_ = re.sub(r'\*', '', input_)
            input_ = re.sub(r'\{New Line\}', ' ', input_)
            input_ = self.tfidf.transform([input_.lower()])
            cosine_sim = cosine_similarity(input_, self.tfidf_matrix)
            i = cosine_sim[0].argmax()
            sim_scores = cosine_sim[0][i]

            true_id = np.array(self.y[['condition_concept_id']].loc[self.index_list[i]].dropna())
            true_name = np.array(self.y[['concept_name']].loc[self.index_list[i]].dropna())
            #true_X = np.array(self.X.iloc[sim_scores[0]])
            concept_id.append(true_id)
            concept_name.append(true_name)
            #mapping_text.append(true_X)
        return concept_id, concept_name#, mapping_text

    def get_similar_simscore(self, statement=None):
        if statement == None: statement = list(self.X)
        concept_id = []
        concept_name = []
        #mapping_text = []
        for input_ in statement:
            input_ = re.sub('-', ' ', input_)
            input_ = re.sub(r'\*', '', input_)
            input_ = re.sub(r'\{New Line\}', ' ', input_)
            input_ = self.tfidf.transform([input_.lower()])
            cosine_sim = cosine_similarity(input_, self.tfidf_matrix)
            i = cosine_sim[0].argmax()
            sim_scores = cosine_sim[0][i]

            true_id = np.array(self.y[['condition_concept_id']].loc[self.index_list[i]].dropna())
            true_name = np.array(self.y[['concept_name']].loc[self.index_list[i]].dropna())
            #true_X = np.array(self.X.iloc[i])
            concept_id.append(true_id)
            concept_name.append(true_name)
            #mapping_text.append(true_X)
        return concept_id, concept_name, sim_scores


    def my_split(self, text, delimiter):
        token = []
        for i in range(len(text)):
            tmp = text[i].split(delimiter)
            for j in range(len(tmp)):
                token.append(tmp[j].strip().lower())
        return token

    def check_if_any(self, statement=None):
        if statement == None: statement = list(self.X)
        concept_id = []
        concept_name = []
        #mapping_text = []

        for input_ in statement:
            input_ = re.sub('-', ' ', input_).lower()
            #input_ = re.sub(r',', '', input_)
            input_ = re.sub(r'\*', '', input_)
            input_ = re.sub(r'\{New Line\}', '', input_)
            input_ = input_.lower()
            out = False
            cont = False

            for tag2 in self.should_not_use:
                if tag2.strip() in input_:
                    for tag3 in self.comment3:
                        if tag3.strip() in input_:
                            cont = True
                    if cont:
                        break
                    concept_id.append([])
                    concept_name.append([])
                    #mapping_text.append(input_)
                    out = True
                    break
            if out:
                continue    # should not use 먼저 검사

            ind = []
            p = 0
            #for not_tag2 in self.if_any_match:
            for not_tag4 in self.X:
                if not_tag4 in input_:
                    add = True
                    for com4 in self.comment4:
                        if com4 == not_tag4:
                            add = False
                    if add:
                        ind.append(self.index_list[p])
                p = p + 1
            output1 = np.array(self.y[['condition_concept_id']].loc[ind].drop_duplicates())
            output2 = np.array(self.y[['concept_name']].loc[ind].drop_duplicates())
            n1 = []
            n2 = []
            for e in range(len(output1)):
                for l in output1[e]:
                    if l != None:
                        n1.append(int(l))
            for e in range(len(output2)):
                for l in output2[e]:
                    if l != None:
                        n2.append(l)

            concept_id.append(self.OrderedSet(n1))
            concept_name.append(self.OrderedSet(n2))
            #mapping_text.append(input_)
        return concept_id, concept_name#, mapping_text
    
